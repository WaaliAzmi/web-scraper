import re
import openpyxl
import logging
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse

# Setup logging
logging.basicConfig(
    filename='scraping_log.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Regular expressions for email and phone number
EMAIL_REGEX = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
PHONE_REGEX = r'(\+\d{1,3})?\s?(\d{3})?[-.\s]?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}'

# Social media platforms to search for
SOCIAL_MEDIA = {
    "facebook": r"https?://(www\.)?facebook.com/[a-zA-Z0-9_.-]+",
    "instagram": r"https?://(www\.)?instagram.com/[a-zA-Z0-9_.-]+",
    "twitter": r"https?://(www\.)?(x|twitter)\.com/[a-zA-Z0-9_.-]+",
    "linkedin": r"https?://(www\.)?linkedin.com/in/[a-zA-Z0-9_.-]+"
}

# Load Excel file with list of websites
def load_websites_from_excel(file_name):
    logging.info(f"Loading Excel file: {file_name}")
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    websites = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            websites.append(row[0])
    logging.info(f"Loaded {len(websites)} websites from the Excel file.")
    return websites, workbook, sheet

# Initialize Selenium WebDriver
def init_webdriver():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--remote-debugging-port=9222")
    chrome_options.add_argument("--window-size=1920,1080")

    logging.info("Initializing WebDriver...")
    try:
        driver = webdriver.Chrome(options=chrome_options)
        logging.info("WebDriver initialized successfully.")
    except WebDriverException as e:
        logging.error(f"Error initializing WebDriver: {e}")
        raise e
    return driver

def scrape_website(driver, url):
    logging.info(f"Scraping website: {url}")
    try:
        driver.get(url)
        # Wait for page content to load
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

        # Scroll down to ensure all dynamic content is loaded
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        WebDriverWait(driver, 5)

        html_content = driver.page_source
        logging.info(f"Successfully scraped content from: {url}")
        return html_content
    except TimeoutException:
        logging.warning(f"Timeout while loading {url}")
    except Exception as e:
        logging.error(f"Error loading {url}: {e}")
    return None

# Prioritize extraction from footer and body for social media links
def extract_info_from_html(html_content):
    logging.info("Extracting information from HTML content.")
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find email addresses
    emails = set(re.findall(EMAIL_REGEX, soup.get_text()))

    # Find phone numbers
    phones = set(re.findall(PHONE_REGEX, soup.get_text()))

    # First, check the footer for social media links
    footer = soup.find('footer')
    social_media_links = {}

    if footer:
        footer_links = footer.find_all('a', href=True)
        for link in footer_links:
            for platform, regex in SOCIAL_MEDIA.items():
                if re.search(regex, link['href']):
                    social_media_links[platform] = link['href']

    # If no links found in footer, check the entire page
    if not social_media_links:
        all_links = soup.find_all('a', href=True)
        for link in all_links:
            for platform, regex in SOCIAL_MEDIA.items():
                if re.search(regex, link['href']):
                    social_media_links[platform] = link['href']

    logging.info(f"Found {len(emails)} emails, {len(phones)} phones, and {len(social_media_links)} social media links.")
    return emails, phones, social_media_links

# Find internal links to crawl
def find_internal_links(soup, base_url):
    links = set()
    for a_tag in soup.find_all("a", href=True):
        href = a_tag['href']
        full_url = urljoin(base_url, href)
        # Filter out external links
        if urlparse(full_url).netloc == urlparse(base_url).netloc:
            links.add(full_url)
    return links

# Update the Excel file with scraped data
def update_excel(sheet, row, emails, phones, social_media_links):
    logging.info(f"Updating Excel sheet for row {row}.")
    email_str = ', '.join(emails) if emails else ''
    
    # Flatten phone numbers (handling tuple issues)
    phone_str = ', '.join([phone if isinstance(phone, str) else ''.join(phone) for phone in phones]) if phones else ''

    # Set the extracted data to corresponding columns
    sheet.cell(row=row, column=2, value=email_str)
    sheet.cell(row=row, column=3, value=phone_str)

    # Social media links
    sheet.cell(row=row, column=4, value=social_media_links.get("facebook", ''))
    sheet.cell(row=row, column=5, value=social_media_links.get("instagram", ''))
    sheet.cell(row=row, column=6, value=social_media_links.get("twitter", ''))
    sheet.cell(row=row, column=7, value=social_media_links.get("linkedin", ''))

# Main function to orchestrate the scraping process
def main(file_name, max_depth=2):
    logging.info("Starting the web scraping process.")
    websites, workbook, sheet = load_websites_from_excel(file_name)
    driver = init_webdriver()

    try:
        for index, website in enumerate(websites, start=2):
            logging.info(f"Processing website {website} (Row {index})")
            visited_urls = set()
            to_visit = {website}
            depth = 0

            all_emails, all_phones, all_social_media_links = set(), set(), {}

            while to_visit and depth < max_depth:
                current_url = to_visit.pop()
                if current_url in visited_urls:
                    continue
                
                html_content = scrape_website(driver, current_url)
                visited_urls.add(current_url)

                if html_content:
                    emails, phones, social_media_links = extract_info_from_html(html_content)
                    all_emails.update(emails)
                    all_phones.update(phones)
                    for platform, link in social_media_links.items():
                        if platform not in all_social_media_links:
                            all_social_media_links[platform] = link

                    soup = BeautifulSoup(html_content, 'html.parser')
                    internal_links = find_internal_links(soup, website)
                    to_visit.update(internal_links)
                
                depth += 1

            update_excel(sheet, index, all_emails, all_phones, all_social_media_links)

        workbook.save(file_name)
        logging.info("Excel file updated successfully.")
    except Exception as e:
        logging.error(f"An error occurred during the process: {e}")
    finally:
        driver.quit()
        logging.info("Web scraping process completed.")

# Usage
if __name__ == "__main__":
    excel_file = "websites_list.xlsx"
    main(excel_file, max_depth=2)
